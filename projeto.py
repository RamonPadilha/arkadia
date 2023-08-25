import openpyxl
import random

# Crie um novo arquivo de planilha
workbook = openpyxl.Workbook()
sheet = workbook.active

# Defina os cabeçalhos das colunas
sheet['A1'] = 'Marca'
sheet['B1'] = 'Modelo'
sheet['C1'] = 'Ano'
sheet['D1'] = 'Preço'

# Lista de marcas e modelos fictícios
marcas = ['Toyota', 'Honda', 'Ford', 'Chevrolet', 'Nissan']
modelos = ['Corolla', 'Civic', 'Camry', 'Accord', 'F-150', 'Mustang', 'Cruze', 'Altima']

# Preencha a planilha com dados aleatórios
for row in range(2, 11):  # Preencha 10 linhas de dados
    sheet.cell(row=row, column=1, value=random.choice(marcas))
    sheet.cell(row=row, column=2, value=random.choice(modelos))
    sheet.cell(row=row, column=3, value=random.randint(2000, 2023))  # Ano entre 2000 e 2023
    sheet.cell(row=row, column=4, value=round(random.uniform(10000, 50000), 2))  # Preço aleatório

# Salve a planilha em um arquivo
workbook.save('dados_carros.xlsx')

print("Planilha 'dados_carros.xlsx' criada com sucesso.")
